# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════
# ROSENI CATERING - TRANSFORMER
# ═══════════════════════════════════════════════════════════════

st.set_page_config(
    page_title="Transformer - Scoro to Buum",
    page_icon="🍴",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Custom CSS - Roseni Catering stiil (must taust, kuldne aktsent)
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@400;600;700&family=Montserrat:wght@300;400;500;600&display=swap');
    
    :root {
        --roseni-black: #0a0a0a;
        --roseni-dark: #1a1a1a;
        --roseni-gold: #c4a265;
        --roseni-gold-light: #d4b88a;
        --roseni-white: #ffffff;
        --roseni-gray: #888888;
    }
    
    .stApp {
        background-color: #0a0a0a;
    }
    
    /* Header area */
    header[data-testid="stHeader"] {
        background-color: #0a0a0a;
    }
    
    /* Logo container */
    .logo-container {
        text-align: center;
        padding: 2rem 0;
        margin-bottom: 1rem;
    }
    
    .logo-container img {
        max-width: 280px;
        height: auto;
    }
    
    /* Main title */
    .main-title {
        font-family: 'Cormorant Garamond', serif;
        color: #c4a265;
        font-size: 2.5rem;
        font-weight: 600;
        text-align: center;
        letter-spacing: 4px;
        margin-bottom: 0.5rem;
    }
    
    .sub-title {
        font-family: 'Montserrat', sans-serif;
        color: #888888;
        font-size: 1rem;
        text-align: center;
        letter-spacing: 2px;
        text-transform: uppercase;
        margin-bottom: 2rem;
    }
    
    /* Gold divider */
    .gold-divider {
        height: 1px;
        background: linear-gradient(90deg, transparent, #c4a265, transparent);
        margin: 2rem auto;
        max-width: 600px;
    }
    
    /* Stat cards */
    .stat-card {
        background: #1a1a1a;
        border: 1px solid #2a2a2a;
        border-radius: 8px;
        padding: 1.5rem;
        text-align: center;
        transition: border-color 0.3s ease;
    }
    
    .stat-card:hover {
        border-color: #c4a265;
    }
    
    .stat-number {
        font-family: 'Cormorant Garamond', serif;
        font-size: 2.5rem;
        font-weight: 700;
        color: #c4a265;
    }
    
    .stat-label {
        font-family: 'Montserrat', sans-serif;
        color: #888888;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 2px;
        margin-top: 0.5rem;
    }
    
    /* Buttons */
    .stButton > button {
        background: transparent;
        color: #c4a265;
        border: 1px solid #c4a265;
        border-radius: 0;
        padding: 0.75rem 2.5rem;
        font-family: 'Montserrat', sans-serif;
        font-weight: 500;
        letter-spacing: 2px;
        text-transform: uppercase;
        transition: all 0.3s ease;
    }
    
    .stButton > button:hover {
        background: #c4a265;
        color: #0a0a0a;
        border-color: #c4a265;
    }
    
    /* File uploader */
    [data-testid="stFileUploader"] {
        background: #1a1a1a;
        border: 1px dashed #333333;
        border-radius: 8px;
        padding: 1rem;
    }
    
    [data-testid="stFileUploader"]:hover {
        border-color: #c4a265;
    }
    
    /* Download button */
    .stDownloadButton > button {
        background: #c4a265;
        color: #0a0a0a;
        border: none;
        border-radius: 0;
        padding: 0.75rem 2.5rem;
        font-family: 'Montserrat', sans-serif;
        font-weight: 600;
        letter-spacing: 2px;
        text-transform: uppercase;
    }
    
    .stDownloadButton > button:hover {
        background: #d4b88a;
        color: #0a0a0a;
    }
    
    /* Dataframe styling */
    .stDataFrame {
        background: #1a1a1a;
        border-radius: 8px;
    }
    
    /* Info/Success/Error messages */
    .stAlert {
        background: #1a1a1a;
        border-radius: 8px;
    }
    
    /* Section headers */
    .section-header {
        font-family: 'Montserrat', sans-serif;
        color: #ffffff;
        font-size: 0.9rem;
        letter-spacing: 2px;
        text-transform: uppercase;
        margin-bottom: 1rem;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Scrollbar */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: #0a0a0a;
    }
    
    ::-webkit-scrollbar-thumb {
        background: #333333;
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: #c4a265;
    }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════
# TRANSFORMATSIOONI FUNKTSIOONID
# ═══════════════════════════════════════════════════════════════

def transform_data(df):
    ""Transformeerib Scoro andmed Buum formaati""
    
    # Filtreeri välja subheading read
    df_filtered = df[df["is_subheading"] != 1].copy()
    
    # Säilita originaaljärjekord
    df_filtered["_original_order"] = range(len(df_filtered))
    
    # Arvuta veerg C (amount x additional_amount)
    df_filtered["calc_c"] = df_filtered["amount"] * df_filtered["additional_amount"]
    
    # Arvuta hindbuumile (hind pärast soodustusi)
    def calc_hindbuumile(row):
        price = row["price"]
        line_disc = row["line_discount_percent"]
        total_disc = row.get("discount", 0)
        
        # Käsitle NaN/tühjad väärtused
        if pd.isna(line_disc):
            line_disc = 0
        if pd.isna(total_disc):
            total_disc = 0
        
        # Kui rea allahindlus on 100%, siis hind on 0
        if line_disc == 100:
            return 0
        
        # ÕIGE VALEM: Korrutamine järjestikku, mitte lahutamine paralleelselt
        # 1. Esmalt rakenda rea-allahindlus (line_discount)
        # 2. Seejärel rakenda üldine allahindlus (discount)
        return price * (1 - line_disc / 100) * (1 - total_disc / 100)
    
    df_filtered["hindbuumile"] = df_filtered.apply(calc_hindbuumile, axis=1)
    
    # Loo väljund tühjade ridadega maksjate vahel
    result_rows = []
    current_payer = None
    columns = ["payer", "product_code", "calc_c", "product_name",
               "amount", "unit", "additional_amount",
               "line_discount_percent", "hindbuumile", "price", "line_sum", "discount"]
    
    for _, row in df_filtered.iterrows():
        if current_payer is not None and current_payer != row["payer"]:
            # --- UUS OSA ALGAB ---
            empty_row = {col: "" for col in columns}
            empty_row["payer"] = "-"
            result_rows.append(empty_row)
            # --- UUS OSA LÕPP ---
            
        current_payer = row["payer"]
        result_rows.append({col: row.get(col, "") for col in columns})
    
    return pd.DataFrame(result_rows)

def create_summary_stats(df):
    """Loob kokkuvõtva statistika"""
    stats = {}
    
    # 1. Filtreerime välja read, mis pole päris andmed (Sinu lisatud "-" read)
    df_clean = df[df["payer"] != "-"].copy()
    
    stats["total_rows"] = len(df_clean)
    stats["unique_payers"] = df_clean["payer"].nunique()
    stats["unique_products"] = df_clean["product_code"].nunique()
    
    try:
        # 2. PÕHILINE PARANDUS: To_numeric koos errors='coerce' 
        # muudab tühjad stringid "" ja muud vead NaN-iks (tühimikuks)
        line_sums = pd.to_numeric(df_clean["line_sum"], errors='coerce')
        
        # 3. Sum() ignoreerib NaN väärtusi automaatselt
        stats["total_sum"] = line_sums.sum()
    except:
        stats["total_sum"] = 0
        
    return stats

def format_excel_output(writer, result_df):
    """Formaadib Excel väljundfaili - paremale joondus ja veeru laiused"""
    workbook = writer.book
    worksheet = writer.sheets["Tulemus"]
    
    # Arvuta iga veeru max laius ja joonda paremale
    for col_idx, column in enumerate(result_df.columns, 1):
        max_length = len(str(column))
        
        # Käi läbi kõik read selles veerus
        for row_idx in range(2, len(result_df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="left")
            cell_value = str(cell.value) if cell.value else ""
            max_length = max(max_length, len(cell_value))
        
        # Lisa natuke ruumi ja sea veeru laius
        adjusted_width = max_length + 2
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Joonda ka päis paremale
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.alignment = Alignment(horizontal="left")

# ═══════════════════════════════════════════════════════════════
# KASUTAJALIIDES
# ═══════════════════════════════════════════════════════════════

# Logo
st.markdown("""
<div class="logo-container">
    <img src="https://rosenievents.ee/wp-content/uploads/2025/03/catering_logo.png" alt="Roseni Catering">
</div>
""", unsafe_allow_html=True)

# Title
st.markdown(<h1 class='main-title'>TRANSFORMER</h1>", unsafe_allow_html=True)"
st.markdown("<p class='sub-title'>Scoro to Buum</p>", unsafe_allow_html=True)
st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)

# File upload
st.markdown("<p class='section-header'>📁 Laadi üles Excel fail</p>", unsafe_allow_html=True)
uploaded_file = st.file_uploader(
    "Vali fail (.xlsx)",
    type=["xlsx", "xls"],
    label_visibility="collapsed"
)

if uploaded_file is not None:
    try:
        # Loe esimene leht automaatselt
        df = pd.read_excel(uploaded_file, sheet_name=0)
        
        st.success(f"✓ Fail laetud: {len(df)} rida")
        
        st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)
        
        # Transform button
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            transform_button = st.button("TRANSFORMEERI", use_container_width=True)
        
        if transform_button:
            with st.spinner("Töötlen..."):
                required_cols = ["is_subheading", "payer", "product_code", "product_name",
                                 "amount", "unit", "additional_amount",
                                 "line_discount_percent", "price", "line_sum"]
                
                missing_cols = [col for col in required_cols if col not in df.columns]
                
                if missing_cols:
                    st.error(f"Puuduvad veerud: {missing_cols}")
                else:
                    result_df = transform_data(df)
                    stats = create_summary_stats(result_df)
                    
                    # Statistics
                    st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)
                    
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.markdown(
                            f"<div class='stat-card'><div class='stat-number'>{stats['total_rows']}</div><div class='stat-label'>Tooterida</div></div>",
                            unsafe_allow_html=True
                        )
                    
                    with col2:
                        st.markdown(
                            f"<div class='stat-card'><div class='stat-number'>{stats['unique_payers']}</div><div class='stat-label'>Maksjat</div></div>",
                            unsafe_allow_html=True
                        )
                    
                    with col3:
                        st.markdown(
                            f"<div class='stat-card'><div class='stat-number'>{stats['unique_products']}</div><div class='stat-label'>Toodet</div></div>",
                            unsafe_allow_html=True
                        )
                    
                    with col4:
                        st.markdown(
                            f"<div class='stat-card'><div class='stat-number'>€{stats['total_sum']:,.0f}</div><div class='stat-label'>Kogusumma</div></div>",
                            unsafe_allow_html=True
                        )
                    
                    st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)
                    
                    # Data table
                    st.markdown("<p class='section-header'>📋 Transformeeritud andmed</p>", unsafe_allow_html=True)
                    st.dataframe(result_df, use_container_width=True, height=400)
                    
                    # Download - formaaditud Excel
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine="openpyxl") as writer:
                        result_df.to_excel(writer, sheet_name="Tulemus", index=False)
                        format_excel_output(writer, result_df)
                    
                    st.markdown("<div class='gold-divider'></div>", unsafe_allow_html=True)
                    
                    col1, col2, col3 = st.columns([1, 2, 1])
                    with col2:
                        st.download_button(
                            label="LAADI ALLA",
                            data=output.getvalue(),
                            file_name="arve boomi.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
    
    except Exception as e:
        st.error(f"Viga: {str(e)}")

else:
    # Welcome message
    st.markdown("""
    <div style="text-align: center; padding: 3rem 0;">
        <p style="color: #888888; font-family: 'Montserrat', sans-serif; font-size: 1rem;">
            Laadi üles Scoro eksport ja transformeeri see Buum formaati
        </p>
    </div>
    """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div style="text-align: center; padding: 3rem 0; margin-top: 2rem;">
    <p style="color: #333333; font-family: 'Montserrat', sans-serif; font-size: 0.75rem; letter-spacing: 2px;">
        ROSENI CATERING
    </p>
</div>
""", unsafe_allow_html=True)
